from pathlib import Path

from openpyxl import load_workbook

from app.core.config import settings
from app.ingestion.loaders.office_common import enforce_text_limit, validate_office_package
from app.models.document import Document, FileType, ParsedUnit, SourceLocator, create_stable_document_id


class XLSXLoader:
    def load(self, file_path: str | Path) -> tuple[Document, list[ParsedUnit]]:
        path = Path(file_path)
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"Unsupported file type: {path.suffix}. Expected .xlsx")
        content = validate_office_package(path, "xl/workbook.xml")
        try:
            workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        except Exception as exc:
            raise ValueError(f"Unable to read XLSX: {path.name}") from exc
        try:
            if len(workbook.worksheets) > settings.xlsx_max_worksheets:
                raise ValueError("Workbook exceeds the configured worksheet limit")
            units = []
            non_empty_cells = 0
            extracted = 0
            for sheet in workbook.worksheets:
                if (
                    sheet.max_row is not None
                    and sheet.max_row > settings.xlsx_max_rows_per_sheet
                ):
                    raise ValueError(
                        f"Worksheet {sheet.title} exceeds the configured row limit"
                    )
                if (
                    sheet.max_column is not None
                    and sheet.max_column > settings.xlsx_max_columns
                ):
                    raise ValueError(
                        f"Worksheet {sheet.title} exceeds the configured column limit"
                    )
                rows = []
                for row_number, cells in enumerate(
                    sheet.iter_rows(max_row=settings.xlsx_max_rows_per_sheet, max_col=settings.xlsx_max_columns),
                    start=1,
                ):
                    values = [self._render_value(cell.value) for cell in cells]
                    if not any(values):
                        continue
                    non_empty_cells += sum(bool(value) for value in values)
                    if non_empty_cells > settings.xlsx_max_non_empty_cells:
                        raise ValueError("Workbook exceeds the configured non-empty cell limit")
                    rows.append((row_number, values))
                if not rows:
                    continue
                headers = [value or f"Column {index + 1}" for index, value in enumerate(rows[0][1])]
                for start in range(1, len(rows), 25):
                    batch = rows[start:start + 25]
                    if not batch:
                        continue
                    lines = [f"Sheet: {sheet.title}"]
                    for _, values in batch:
                        lines.append("\n".join(
                            f"{header}: {value}" for header, value in zip(headers, values) if value
                        ))
                    rendered = "\n\n".join(lines)
                    extracted = enforce_text_limit(extracted, rendered)
                    row_start, row_end = batch[0][0], batch[-1][0]
                    units.append(ParsedUnit(
                        text=rendered,
                        section=sheet.title,
                        source_locator=SourceLocator(
                            type="sheet_rows",
                            label=f"{sheet.title} · rows {row_start}–{row_end}",
                        ),
                    ))
            if not units:
                raise ValueError("XLSX contains no extractable data rows")
            return Document(
                id=create_stable_document_id(content, FileType.XLSX),
                filename=path.name,
                file_type=FileType.XLSX,
            ), units
        finally:
            workbook.close()

    @staticmethod
    def _render_value(value) -> str:
        if value is None:
            return ""
        if isinstance(value, str) and value.startswith("="):
            return f"[Formula: {value}]"
        return str(value).strip()
